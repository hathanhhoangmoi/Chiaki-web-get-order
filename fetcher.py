from curl_cffi.requests import AsyncSession
import io
import json
import re
import urllib.parse
from datetime import datetime, timedelta, timezone
from sqlalchemy.orm import Session
from models import Order, ShopMeta
from shops_config import SELLER_ID, SELLER_TOKEN
from openpyxl import load_workbook

# ✅ THÊM: URL Cloudflare Worker proxy
CHIAKI_PROXY = "https://chiakicl.hathanhhoang-edu.workers.dev"


def build_api_url(shop_id: str) -> str:
    VN_TZ = timezone(timedelta(hours=7))
    today = datetime.now(VN_TZ)
    since = today - timedelta(days=14)
    def fmt(d): return d.strftime("%d/%m/%Y").replace("/", "%2F")
    range_str = f"{fmt(since)}%20-%20{fmt(today)}"

    # ✅ SỬA: build direct_url trước, sau đó wrap qua proxy
    direct_url = (
        f"https://api.chiaki.vn/api/{shop_id}/export-excel-order"
        f"?source=seller&page_index=1&page_size=500&status=receive_wating"
        f"&range_date={range_str}"
        f"&date_type=created_at&order=create-desc"
        f"&Seller_id={SELLER_ID}&Seller_token={SELLER_TOKEN}"
    )
    return f"{CHIAKI_PROXY}?url={urllib.parse.quote(direct_url, safe='')}"


async def fetch_shop_name(shop_url: str) -> str:
    """Giữ lại cho /api/test-shopname, không dùng trong sync nữa"""
    try:
        async with httpx.AsyncClient(
            timeout=10,
            follow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0"}
        ) as client:
            res = await client.get(shop_url)
            if not res.is_success:
                return shop_url

            patterns = [
                r'<span[^>]*class=["\']store-title["\'][^>]*>(.*?)</span>',
                r'class=["\']store-title["\'][^>]*>(.*?)<',
                r'store-title["\']>(.*?)<',
            ]
            for pattern in patterns:
                m = re.search(pattern, res.text, re.IGNORECASE | re.DOTALL)
                if m:
                    name = m.group(1).strip()
                    if name:
                        print(f"[fetch_name] {shop_url} → {name}")
                        return name

            print(f"[fetch_name] Không tìm thấy tên: {shop_url}")
            return shop_url
    except Exception as e:
        print(f"[fetch_name] Error {shop_url}: {e}")
        return shop_url


def parse_excel(content: bytes, shop_id: str, shop_name: str) -> list[dict]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        headers = [str(c).strip() if c else "" for c in rows[0]]
        print(f"[parse] headers: {headers}")

        def find_col(keywords):
            for kw in keywords:
                for i, h in enumerate(headers):
                    if kw.lower() in h.lower():
                        return i
            return None

        col_code     = find_col(["mã đơn hàng", "mã đơn", "order_id"])
        col_customer = find_col(["người đặt hàng", "người đặt", "đặt hàng", "khách"])
        col_buyer    = find_col(["tên người nhận", "người nhận", "buyer"])
        col_phone    = find_col(["sđt", "điện thoại", "phone", "số điện thoại"])
        col_address  = find_col(["địa chỉ", "address"])
        col_product  = find_col(["tên sản phẩm", "ten san pham", "tên hàng", "product name", "product"])
        col_qty      = find_col(["số lượng", "quantity", "qty", "sl"])
        col_total    = find_col(["tổng tiền", "tổng", "total", "amount"])
        col_status   = find_col(["trạng thái", "status"])
        col_date     = find_col(["thời gian đặt hàng", "thời gian đặt", "thời gian", "ngày đặt", "ngày tạo", "ngày", "date", "time"])

        def val(row, idx):
            if idx is None or idx >= len(row): return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        orders = []
        for i, row in enumerate(rows[1:]):
            code = f"{shop_id}_{val(row, col_code)}" if val(row, col_code) else f"{shop_id}_{i}"
            if not code or code == "None": continue
            try:
                qty   = int(float(val(row, col_qty)))  if val(row, col_qty)   else 0
                total = float(val(row, col_total))      if val(row, col_total) else 0.0
            except:
                qty, total = 0, 0.0
            orders.append({
                "order_code":    code,
                "shop_id":       shop_id,
                "shop_name":     shop_name,
                "buyer_name":    val(row, col_buyer),
                "customer_name": val(row, col_customer),
                "phone":         val(row, col_phone),
                "address":       val(row, col_address),
                "product":       val(row, col_product),
                "quantity":      qty,
                "total":         total,
                "status":        val(row, col_status),
                "order_date":    val(row, col_date),
                "raw_data":      json.dumps(
                    dict(zip(headers, [str(c) for c in row])),
                    ensure_ascii=False
                ),
            })
        return orders
    except Exception as e:
        print(f"[parse_excel] Error shop {shop_id}: {e}")
        return []


async def sync_shop(shop_id: str, shop_url: str, shop_name: str, db: Session) -> int:
    url = build_api_url(shop_id)
    print(f"[fetch] {shop_id} → {url}")

    try:
        # ✅ SỬA: dùng httpx thay vì curl_cffi vì request đến Workers (không bị Cloudflare chặn)
        import httpx
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            res = await client.get(url)
            print(f"[fetch] {shop_id} status={res.status_code} size={len(res.content)} bytes")
            if res.status_code != 200:
                print(f"[fetch] {shop_id} FAILED: {res.text[:200]}")
                return 0
            content = res.content
    except Exception as e:
        print(f"[fetch] {shop_id} exception: {e}")
        return 0

    orders = parse_excel(content, shop_id, shop_name)
    print(f"[parse] {shop_id} → {len(orders)} đơn đọc được từ Excel")

    deleted = db.query(Order).filter(Order.shop_id == shop_id).delete()
    print(f"[delete] {shop_id} → đã xoá {deleted} đơn cũ")

    for o in orders:
        db.add(Order(**o))
    db.commit()

    meta = db.query(ShopMeta).filter(ShopMeta.shop_id == shop_id).first()
    if meta:
        meta.shop_name   = shop_name
        meta.last_sync   = datetime.now(timezone(timedelta(hours=7)))
        meta.order_count = len(orders)
    else:
        db.add(ShopMeta(
            shop_id=shop_id, shop_name=shop_name,
            shop_url=shop_url, last_sync=datetime.now(timezone(timedelta(hours=7))),
            order_count=len(orders)
        ))
    db.commit()

    print(f"[sync] {shop_name} ({shop_id}): đã thay thế → {len(orders)} đơn")
    return len(orders)
