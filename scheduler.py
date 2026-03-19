import asyncio
import httpx
import openpyxl
import urllib.parse
from io import BytesIO
from datetime import datetime, timedelta
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from database import SessionLocal
from fetcher import sync_shop
from shops_config import get_shops_map, SELLER_ID, SELLER_TOKEN
from models import (
    Order, DeliveringOrder, FinishedOrder,
    ReturnedOrder, RevenueCache, RevenueNetCache, ShopMeta
)
PROXY_URL = "https://chiaki-proxy.hathanhhoang-edu.workers.dev"
scheduler = AsyncIOScheduler(timezone="Asia/Ho_Chi_Minh")

# ─────────────────────────────────────────
# HEADERS GIẢ LẬP BROWSER (tránh Cloudflare)
# ─────────────────────────────────────────
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "vi-VN,vi;q=0.9",
    "Referer": "https://seller.chiaki.vn/",
    "Origin": "https://seller.chiaki.vn",
}


# ─────────────────────────────────────────
# HELPER: BUILD DATE RANGE
# ─────────────────────────────────────────
def build_date_range(days=30):
    end = datetime.now()
    start = end - timedelta(days=days)
    date_range = f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"
    return date_range, urllib.parse.quote(date_range)


# ─────────────────────────────────────────
# HELPER: FETCH EXCEL ĐƠN HÀNG TỪNG SHOP
# ─────────────────────────────────────────

# ─────────────────────────────────────────
# SYNC ĐƠN HÀNG THEO STATUS VÀO DB
# ─────────────────────────────────────────
async def sync_status_orders(status: str, ModelClass, days=30):
    date_range, encoded_range = build_date_range(days)
    shops = get_shops_map()
    semaphore = asyncio.Semaphore(5)

    async def fetch_one(client, sid, sname):
        async with semaphore:
            return await fetch_excel_orders(client, sid, sname, status, encoded_range)

    async with httpx.AsyncClient(timeout=30) as client:
        tasks = [fetch_one(client, sid, val[1]) for sid, val in shops.items()]
        results = await asyncio.gather(*tasks)

    all_orders = [o for shop_orders in results for o in shop_orders]

    db = SessionLocal()
    try:
        db.query(ModelClass).delete()
        for o in all_orders:
            db.add(ModelClass(**o))
        db.commit()
        print(f"  ✅ [{status}] Lưu {len(all_orders)} đơn vào DB")
    except Exception as e:
        db.rollback()
        print(f"  ❌ [{status}] DB error: {e}")
    finally:
        db.close()

    return len(all_orders)

# FETCH EXCEL ĐƠN HÀNG TỪNG SHOP
# ─────────────────────────────────────────
async def fetch_excel_orders(client, sid, shop_name, status, encoded_range, pagesize=500):
    chiaki_url = (
        f"https://api.chiaki.vn/api/{sid}/export-excel-order"
        f"?source=seller&pageIndex=1&pageSize={pagesize}"
        f"&status={status}&rangeDate={encoded_range}"
        f"&dateType=createdat&order=create-desc"
        f"&SellerId={SELLER_ID}&SellerToken={SELLER_TOKEN}"
    )
    url = f"{PROXY_URL}?url={urllib.parse.quote(chiaki_url, safe='')}"

    try:
        resp = await client.get(url)
        if resp.status_code != 200:
            print(f"  [{status}] {sid} HTTP {resp.status_code}")
            return []
        ct = resp.headers.get("content-type", "")
        if "html" in ct or len(resp.content) < 100:
            print(f"  [{status}] {sid} FAILED (bị chặn hoặc rỗng)")
            return []

        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        headers_map, header_row = {}, None

        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and any(cell and "mã đơn" in str(cell).lower() for cell in row):
                header_row = i
                for j, cell in enumerate(row):
                    if cell:
                        headers_map[str(cell).strip().lower()] = j
                break

        if header_row is None:
            return []

        def fc(kw):
            for k, v in headers_map.items():
                if kw.lower() in k:
                    return v
            return None

        col_code    = fc("mã đơn")
        col_date    = fc("ngày tạo") or fc("ngày tạo")
        col_name    = fc("tên người nhận") or fc("tên khách")
        col_phone   = fc("số điện thoại") or fc("điện thoại")
        col_address = fc("địa chỉ")
        col_product = fc("tên sản phẩm") or fc("sản phẩm")
        col_total   = fc("tổng tiền") or fc("giá trị")

        orders = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(c is None for c in row):
                continue

            def gv(col):
                return str(row[col]).strip() if col is not None and row[col] is not None else ""

            orders.append({
                "order_code":    gv(col_code),
                "shop_id":       sid,
                "shop_name":     shop_name,
                "buyer_name":    gv(col_name),
                "customer_name": gv(col_name),
                "phone":         gv(col_phone),
                "address":       gv(col_address),
                "product":       gv(col_product),
                "quantity":      "",
                "total":         gv(col_total),
                "order_date":    gv(col_date),
            })

        print(f"  [{status}] {sid} → {len(orders)} đơn")
        return orders

    except Exception as e:
        print(f"  [{status}] {sid} lỗi: {e}")
        return []


# ─────────────────────────────────────────
# SYNC DOANH THU GỘP (30 ngày)
# ─────────────────────────────────────────
async def sync_revenue(days=30):
    date_range, encoded_range = build_date_range(days)
    shops = get_shops_map()
    semaphore = asyncio.Semaphore(5)

    async def fetch_one(client, sid, sname):
        async with semaphore:
            chiaki_url = (
                f"https://api.chiaki.vn/api/{sid}/export-excel-summary-amount-order"
                f"?source=seller&pageIndex=1&pageSize=500"
                f"&status=all&rangeDate={encoded_range}"
                f"&dateType=createdat&order=create-desc"
                f"&SellerId={SELLER_ID}&SellerToken={SELLER_TOKEN}"
            )
            url = f"{PROXY_URL}?url={urllib.parse.quote(chiaki_url, safe='')}"
            try:
                resp = await client.get(url)
                if resp.status_code != 200:
                    print(f"  [revenue] {sid} HTTP {resp.status_code}")
                    return None
                ct = resp.headers.get("content-type", "")
                if "html" in ct or len(resp.content) < 100:
                    print(f"  [revenue] {sid} FAILED (bị chặn hoặc rỗng)")
                    return None

                wb = openpyxl.load_workbook(BytesIO(resp.content))
                ws = wb.active

                def get_val(key_text):
                    for row in ws.iter_rows():
                        for i, cell in enumerate(row):
                            if cell.value and key_text.lower() in str(cell.value).lower():
                                if i + 1 < len(row) and row[i + 1].value is not None:
                                    return row[i + 1].value
                    return None

                def to_float(val):
                    if val is None:
                        return 0
                    try:
                        return float(val)
                    except:
                        return 0

                gia_goc         = to_float(get_val("Giá gốc"))
                hoan_lai        = to_float(get_val("Số tiền hoàn lại"))
                tro_gia         = to_float(get_val("Sản phẩm được trợ giá"))
                mau_u_dai       = to_float(get_val("Mẫu ưu đãi do Người Bán chịu"))
                doanh_thu_gop   = gia_goc + hoan_lai + tro_gia + mau_u_dai

                phi_co_dinh     = to_float(get_val("Phí cố định"))
                phi_dich_vu     = to_float(get_val("Phí Dịch Vụ"))
                phi_tt          = to_float(get_val("Phí thanh toán"))
                phi_qc          = to_float(get_val("Phí quảng cáo"))
                thue_gtgt       = to_float(get_val("Thuế GTGT"))
                thue_tncn       = to_float(get_val("Thuế TNCN"))
                tong_khau_tru   = phi_co_dinh + phi_dich_vu + phi_tt + phi_qc + thue_gtgt + thue_tncn
                doanh_thu_thuan = doanh_thu_gop - tong_khau_tru

                print(f"  [revenue] {sid} → {doanh_thu_gop:,.0f}đ")
                return {
                    "shop_id":          sid,
                    "shop_name":        sname,
                    "ten_shop":         str(get_val("Người Bán") or sname),
                    "chu_tk":           str(get_val("Tên chủ tài khoản") or ""),
                    "ngan_hang":        str(get_val("Tên ngân hàng") or ""),
                    "stk":              str(get_val("Tài khoản ngân hàng") or ""),
                    "doanh_thu_gop":    doanh_thu_gop,
                    "tong_khau_tru":    tong_khau_tru,
                    "doanh_thu_thuan":  doanh_thu_thuan,
                    "date_range":       date_range,
                }
            except Exception as e:
                print(f"  [revenue] {sid} lỗi: {e}")
                return None

    async with httpx.AsyncClient(timeout=30) as client:
        tasks = [fetch_one(client, sid, val[1]) for sid, val in shops.items()]
        results = await asyncio.gather(*tasks)

    valid = [r for r in results if r is not None]
    db = SessionLocal()
    try:
        db.query(RevenueCache).delete()
        for r in valid:
            db.add(RevenueCache(**r))
        db.commit()
        print(f"  ✅ [revenue] Lưu {len(valid)} shop vào DB")
    except Exception as e:
        db.rollback()
        print(f"  ❌ [revenue] DB error: {e}")
    finally:
        db.close()


# ─────────────────────────────────────────
# SYNC DOANH THU THUẦN (14 ngày)
# ─────────────────────────────────────────
async def sync_revenue_net(days=14):
    date_range, encoded_range = build_date_range(days)
    shops = get_shops_map()
    semaphore = asyncio.Semaphore(5)

    async def fetch_one(client, sid, sname):
        async with semaphore:
            chiaki_url = (
                f"https://api.chiaki.vn/api/{sid}/export-excel-order"
                f"?source=seller&pageIndex=1&pageSize=500"
                f"&status=finished&rangeDate={encoded_range}"
                f"&dateType=createdat&order=create-desc"
                f"&SellerId={SELLER_ID}&SellerToken={SELLER_TOKEN}"
            )
            url = f"{PROXY_URL}?url={urllib.parse.quote(chiaki_url, safe='')}"
            try:
                resp = await client.get(url)
                if resp.status_code != 200:
                    print(f"  [revenue_net] {sid} HTTP {resp.status_code}")
                    return None
                ct = resp.headers.get("content-type", "")
                if "html" in ct or len(resp.content) < 100:
                    print(f"  [revenue_net] {sid} FAILED")
                    return None

                wb = openpyxl.load_workbook(BytesIO(resp.content))
                ws = wb.active
                headers_map, header_row = {}, None

                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row and any(cell and "tổng tiền" in str(cell).lower() for cell in row):
                        header_row = i
                        for j, cell in enumerate(row):
                            if cell:
                                headers_map[str(cell).strip().lower()] = j
                        break

                if header_row is None:
                    return None

                def find_col(keyword):
                    for k, v in headers_map.items():
                        if keyword.lower() in k:
                            return v
                    return None

                col_total   = find_col("tổng tiền")
                col_phuphi  = find_col("phụ phí")
                col_doisoat = find_col("ngày đối soát")

                if col_total is None:
                    return None

                tong_tien_sum, phu_phi_sum, row_count = 0, 0, 0
                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    if not row or all(c is None for c in row):
                        continue
                    doi_soat = row[col_doisoat] if col_doisoat is not None else None
                    if doi_soat is not None and str(doi_soat).strip() != "":
                        continue
                    try:
                        tong_tien_sum += float(row[col_total] or 0)
                    except:
                        pass
                    try:
                        phu_phi_sum += float(row[col_phuphi] or 0) if col_phuphi is not None else 0
                    except:
                        pass
                    row_count += 1

                print(f"  [revenue_net] {sid} → {row_count} đơn, {tong_tien_sum:,.0f}đ")
                return {
                    "shop_id":          sid,
                    "shop_name":        sname,
                    "row_count":        row_count,
                    "tong_tien":        tong_tien_sum,
                    "phu_phi":          phu_phi_sum,
                    "doanh_thu_thuan":  tong_tien_sum - phu_phi_sum,
                    "date_range":       date_range,
                }
            except Exception as e:
                print(f"  [revenue_net] {sid} lỗi: {e}")
                return None

    async with httpx.AsyncClient(timeout=30) as client:
        tasks = [fetch_one(client, sid, val[1]) for sid, val in shops.items()]
        results = await asyncio.gather(*tasks)

    valid = [r for r in results if r is not None]
    db = SessionLocal()
    try:
        db.query(RevenueNetCache).delete()
        for r in valid:
            db.add(RevenueNetCache(**r))
        db.commit()
        print(f"  ✅ [revenue_net] Lưu {len(valid)} shop vào DB")
    except Exception as e:
        db.rollback()
        print(f"  ❌ [revenue_net] DB error: {e}")
    finally:
        db.close()
async def sync_all_shops():
    """Sync đơn chờ lấy hàng vào bảng orders (dùng fetcher.py cũ)"""
    print("[scheduler] Bắt đầu quét toàn bộ gian hàng...")
    shops = get_shops_map()

    async def sync_one(sid, url, name):
        db = SessionLocal()
        try:
            return await sync_shop(sid, url, name, db)
        finally:
            db.close()

    tasks = [sync_one(sid, url, name) for sid, (url, name) in shops.items()]
    results = await asyncio.gather(*tasks, return_exceptions=True)
    total_new = sum(r for r in results if isinstance(r, int))
    print(f"[scheduler] ✅ Hoàn tất chuẩn bị hàng: +{total_new} đơn mới")


# ─────────────────────────────────────────
# SYNC TẤT CẢ (chạy mỗi 1 tiếng)
# ─────────────────────────────────────────
async def sync_all():
    print("\n" + "="*50)
    print(f"🔄 BẮT ĐẦU SYNC - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("="*50)

    await asyncio.gather(
        sync_all_shops(),                                        # Chờ lấy hàng
        sync_status_orders("delivering", DeliveringOrder, 30),  # Đang giao
        sync_status_orders("finished",   FinishedOrder,   30),  # Đã giao
        sync_status_orders("returned",   ReturnedOrder,   30),  # Hoàn hàng
        sync_revenue(days=30),                                   # Doanh thu gộp
        sync_revenue_net(days=14),                               # Doanh thu thuần
    )

    print("="*50)
    print(f"✅ SYNC XONG - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("="*50 + "\n")


# ─────────────────────────────────────────
# KHỞI ĐỘNG SCHEDULER
# ─────────────────────────────────────────
def start_scheduler():
    scheduler.add_job(
        sync_all,
        trigger="interval",
        hours=1,
        id="sync_all",
        replace_existing=True,
        next_run_time=datetime.now()  # Chạy ngay khi app khởi động
    )
    scheduler.start()
    print("⏰ Scheduler đã khởi động - sync mỗi 1 tiếng")
