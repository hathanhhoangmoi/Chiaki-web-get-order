import io
import json as _json
import re
from datetime import datetime, timedelta

import httpx
from fastapi import Depends, FastAPI, File, Form, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy import desc, func, or_
from sqlalchemy.orm import Session

from database import engine, get_db, migrate
from fetcher import sync_shop, parse_excel
from models import Base, ExternalOrderConfig, ExternalOrderTracking, Order, ShopMeta
from shops_config import BLOCKED_SHOPS, SELLER_ID, SELLER_TOKEN, SHOP_NAME_MAP, get_shops_map

# ── Key management cho order-info ─────────────────────────
VALID_KEYS = {
    "HATHANHHOANG": 0,
    "PHONE-KEY-PHUONG2000": 0,
}
KEY_LIMIT = 10
# Lưu lịch sử tra cứu: {key: [{"order_code": ..., "time": ...}]}
KEY_HISTORY: dict = {k: [] for k in VALID_KEYS}
# Lưu lịch sử đăng nhập: {key: [{"event": "login/logout", "time": ...}]}
LOGIN_HISTORY: list = []  # [{key, event, time}]
# Database setup
Base.metadata.create_all(bind=engine)
migrate()
UNLIMITED_KEYS = {"HATHANHHOANG", "PHONE-KEY-PHUONG2000"}

app = FastAPI(title="Chiaki Order Dashboard")
app.mount("/static", StaticFiles(directory="static"), name="static")

# ── Helper functions ───────────────────────────────────────
FULL_ACCESS_IDS = {"Chang2000"}
SENSITIVE_SHOPS = {"4647", "4732"}
SENSITIVE_TOTAL_THRESHOLD = 2_500_000
PHONE_KEY_PHUONG_ALLOWED_SHOPS = {"4917", "4940", "5096", "5125", "5114"}

def is_full_access_user(user_id: str) -> bool:
    return str(user_id or "").strip() in FULL_ACCESS_IDS

def should_hide_order(order, user_id: str) -> bool:
    if not order or is_full_access_user(user_id):
        return False
    shop_id = str(getattr(order, "shop_id", "") or "").strip()
    total = getattr(order, "total", 0) or 0
    try:
        total = float(total)
    except Exception:
        total = 0
    return shop_id in SENSITIVE_SHOPS or total >= SENSITIVE_TOTAL_THRESHOLD

def serialize_order(o):
    return {
        "order_code":    o.order_code,
        "order_date":    o.order_date,
        "shop_id":       o.shop_id,
        "shop_name":     o.shop_name,
        "buyer_name":    o.buyer_name,
        "customer_name": o.customer_name,
        "address":       o.address,
        "product":       o.product,
        "quantity":      o.quantity,
        "total":         o.total,
        "status":        o.status,
        "fetched_at":    o.fetched_at.isoformat() if o.fetched_at else None,
        "restricted":    False,
    }


def can_manage_external_orders(user_id: str) -> bool:
    return str(user_id or "").strip() == "Chang2000"


def serialize_external_order(o: ExternalOrderTracking):
    return {
        "code": o.order_code,
        "cod": int(o.cod_amount or 0),
        "status": o.status or "unknown",
        "is_paid": bool(o.is_paid),
        "updated_at": o.updated_at.isoformat() if o.updated_at else None,
    }


def get_external_order_config(db: Session) -> ExternalOrderConfig:
    config = db.query(ExternalOrderConfig).filter(ExternalOrderConfig.id == 1).first()
    if not config:
        config = ExternalOrderConfig(id=1, fee_items_json="[]")
        db.add(config)
        db.commit()
        db.refresh(config)
    return config


def serialize_external_order_config(config: ExternalOrderConfig):
    fee_items = []
    try:
        raw = config.fee_items_json or "[]"
        parsed = _json.loads(raw)
        if isinstance(parsed, list):
            fee_items = parsed
    except Exception:
        fee_items = []
    return {
        "fee_items": fee_items,
        "updated_at": config.updated_at.isoformat() if config.updated_at else None,
    }

# ── API Endpoints ──────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def root():
    with open("static/index.html", encoding="utf-8") as f:
        return f.read()

@app.get("/api/summary")
def get_summary(db: Session = Depends(get_db)):
    shops = db.query(ShopMeta).all()
    total = db.query(Order).count()
    return {
        "total_orders": total,
        "total_shops": len(shops),
        "shops": [
            {
                "shop_id": s.shop_id,
                "shop_name": s.shop_name,
                "order_count": s.order_count,
                "last_sync": s.last_sync.isoformat() if s.last_sync else None,
            }
            for s in shops
        ]
    }

@app.get("/api/orders")
def get_orders(
    request: Request,
    shop_id: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(200, le=200),
    sort: str = Query("default"),
    db: Session = Depends(get_db)
):
    user_id = request.headers.get('X-User-ID', '')

    if shop_id and shop_id in BLOCKED_SHOPS and not is_full_access_user(user_id):
        return {
            "total": 0, "page": page, "data": [],
            "blocked": True,
            "message": "Shop này bị chặn trích xuất đơn hàng"
        }

    q = db.query(Order)
    if shop_id:
        q = q.filter(Order.shop_id == shop_id)
    if not is_full_access_user(user_id):
        q = q.filter(~Order.shop_id.in_(SENSITIVE_SHOPS)).filter(
            or_(Order.total == None, Order.total < SENSITIVE_TOTAL_THRESHOLD)
        )
    total = q.count()

    # Sắp xếp
    if sort == "total_desc":
        q = q.order_by(desc(Order.total))
    elif sort == "total_asc":
        q = q.order_by(Order.total)
    elif sort == "date_desc":
        q = q.order_by(desc(Order.order_date))
    elif sort == "date_asc":
        q = q.order_by(Order.order_date)
    else:
        q = q.order_by(desc(Order.fetched_at))

    orders = q.offset((page - 1) * limit).limit(limit).all()

    return {
        "total": total,
        "page": page,
        "data": [serialize_order(o) for o in orders]
    }

@app.get("/api/orders/search-products")
def search_orders_by_product(
    request: Request,
    q: str = Query(..., min_length=1),
    limit: int = Query(200, ge=1, le=500),
    db: Session = Depends(get_db)
):
    user_id = request.headers.get('X-User-ID', '')
    normalized_q = q.strip().lower()
    tokens = [token.strip().lower() for token in normalized_q.split() if token.strip()]
    if not normalized_q or not tokens:
        return {"total": 0, "data": []}

    filters = [func.lower(Order.product).contains(normalized_q)]
    filters.extend(func.lower(Order.product).contains(token) for token in tokens)

    q_orders = db.query(Order).filter(or_(*filters))
    if not is_full_access_user(user_id):
        q_orders = q_orders.filter(~Order.shop_id.in_(SENSITIVE_SHOPS)).filter(
            or_(Order.total == None, Order.total < SENSITIVE_TOTAL_THRESHOLD)
        )
    orders = q_orders.order_by(desc(Order.order_date), desc(Order.fetched_at)).limit(limit).all()

    return {
        "total": len(orders),
        "data": [serialize_order(o) for o in orders]
    }

@app.post("/api/update-shopname")
def update_shopname(body: dict, db: Session = Depends(get_db)):
    shop_id = body.get("shop_id")
    shop_name = body.get("shop_name")
    if not shop_id or not shop_name:
        return {"ok": False}
    meta = db.query(ShopMeta).filter(ShopMeta.shop_id == shop_id).first()
    if meta:
        meta.shop_name = shop_name
        db.query(Order).filter(Order.shop_id == shop_id).update({"shop_name": shop_name})
        db.commit()
    return {"ok": True, "shop_id": shop_id, "shop_name": shop_name}

@app.get("/api/orders/hanoi")
async def get_hanoi_orders(request: Request, db: Session = Depends(get_db)):
    user_id = request.headers.get('X-User-ID', '')
    keywords = ["hà nội", "ha noi", " hn", "hanoi", "Hà Nội"]
    filters = [func.lower(Order.address).contains(kw.lower()) for kw in keywords]
    q = db.query(Order).filter(or_(*filters))
    if not is_full_access_user(user_id):
        q = q.filter(~Order.shop_id.in_(SENSITIVE_SHOPS)).filter(
            or_(Order.total == None, Order.total < SENSITIVE_TOTAL_THRESHOLD)
        )
    orders = q.order_by(Order.order_date.desc()).all()
    return [serialize_order(o) for o in orders]


@app.get("/api/orders/nuochoa")
async def get_nuochoa_orders(request: Request, db: Session = Depends(get_db)):
    user_id = request.headers.get('X-User-ID', '')
    keywords = ["nước hoa", "nuoc hoa", "nươc hoa", "nước  hoa"]
    filters = [func.lower(Order.product).contains(kw.lower()) for kw in keywords]
    q = db.query(Order).filter(or_(*filters))
    if not is_full_access_user(user_id):
        q = q.filter(~Order.shop_id.in_(SENSITIVE_SHOPS)).filter(
            or_(Order.total == None, Order.total < SENSITIVE_TOTAL_THRESHOLD)
        )
    orders = q.order_by(Order.order_date.desc()).all()
    return [serialize_order(o) for o in orders]

@app.get("/api/chart-data")
def get_chart_data(db: Session = Depends(get_db)):
    date_col = func.substr(Order.order_date, 1, 10)
    by_date = (
        db.query(date_col, func.count(Order.id))
        .filter(Order.order_date != None, Order.order_date != '')
        .group_by(date_col)
        .order_by(date_col)
        .all()
    )
    by_shop = (
        db.query(Order.shop_name, func.count(Order.id))
        .filter(Order.shop_name != None, Order.shop_name != '')
        .group_by(Order.shop_name)
        .order_by(func.count(Order.id).desc())
        .all()
    )
    return {
        "by_date": [{"date": d, "count": c} for d, c in by_date if d],
        "by_shop": [{"shop": s, "count": c} for s, c in by_shop if s],
    }


@app.get("/api/external-orders")
def get_external_orders(db: Session = Depends(get_db)):
    rows = db.query(ExternalOrderTracking).order_by(desc(ExternalOrderTracking.updated_at), ExternalOrderTracking.order_code).all()
    config = get_external_order_config(db)
    return {
        "total": len(rows),
        "items": [serialize_external_order(row) for row in rows],
        "config": serialize_external_order_config(config),
    }


@app.post("/api/external-orders")
def save_external_orders(request: Request, body: dict, db: Session = Depends(get_db)):
    user_id = request.headers.get("X-User-ID", "")
    if not can_manage_external_orders(user_id):
        return JSONResponse({"error": "Không có quyền chỉnh cấu hình."}, status_code=403)

    items = body.get("items")
    if not isinstance(items, list):
        return JSONResponse({"error": "Dữ liệu không hợp lệ."}, status_code=422)

    normalized = []
    seen = set()
    valid_statuses = {"unknown", "in_transit", "delivered", "returned"}
    for item in items:
        if not isinstance(item, dict):
            continue
        code = str(item.get("code", "")).strip().upper()
        if not code or code in seen:
            continue
        seen.add(code)
        try:
            cod = int(item.get("cod") or 0)
        except Exception:
            cod = 0
        status = str(item.get("status") or "unknown").strip()
        if status not in valid_statuses:
            status = "unknown"
        is_paid = 1 if item.get("is_paid") else 0
        normalized.append({"code": code, "cod": cod, "status": status, "is_paid": is_paid})

    fee_items_in = body.get("fee_items")
    fee_items = []
    if isinstance(fee_items_in, list):
        for item in fee_items_in:
            if not isinstance(item, dict):
                continue
            content = str(item.get("content", "") or "").strip()
            try:
                amount = int(item.get("amount") or 0)
            except Exception:
                amount = 0
            collected = bool(item.get("collected"))
            if not content and amount == 0:
                continue
            fee_items.append({
                "content": content,
                "amount": amount,
                "collected": collected,
            })

    db.query(ExternalOrderTracking).delete()
    for item in normalized:
        db.add(ExternalOrderTracking(
            order_code=item["code"],
            cod_amount=item["cod"],
            status=item["status"],
            is_paid=item["is_paid"],
            updated_at=datetime.now(),
        ))
    config = get_external_order_config(db)
    config.fee_items_json = _json.dumps(fee_items, ensure_ascii=False)
    config.updated_at = datetime.now()
    db.commit()

    rows = db.query(ExternalOrderTracking).order_by(desc(ExternalOrderTracking.updated_at), ExternalOrderTracking.order_code).all()
    return {
        "ok": True,
        "total": len(rows),
        "items": [serialize_external_order(row) for row in rows],
        "config": serialize_external_order_config(config),
    }


@app.post("/api/external-orders/status")
def update_external_order_status(request: Request, body: dict, db: Session = Depends(get_db)):
    user_id = request.headers.get("X-User-ID", "").strip()
    if not can_manage_external_orders(user_id):
        return JSONResponse({"error": "Không có quyền cập nhật trạng thái."}, status_code=403)

    code = str(body.get("code", "")).strip().upper()
    status = str(body.get("status", "unknown")).strip()
    valid_statuses = {"unknown", "in_transit", "delivered", "returned"}
    if not code or status not in valid_statuses:
        return JSONResponse({"error": "Dữ liệu không hợp lệ."}, status_code=422)

    row = db.query(ExternalOrderTracking).filter(ExternalOrderTracking.order_code == code).first()
    if not row:
        return JSONResponse({"error": "Không tìm thấy đơn ngoại sàn."}, status_code=404)

    row.status = status
    row.updated_at = datetime.now()
    db.commit()

    return {
        "ok": True,
        "item": serialize_external_order(row),
    }


@app.post("/api/external-orders/payment")
def update_external_order_payment(request: Request, body: dict, db: Session = Depends(get_db)):
    user_id = request.headers.get("X-User-ID", "").strip()
    if not can_manage_external_orders(user_id):
        return JSONResponse({"error": "Không có quyền cập nhật thanh toán."}, status_code=403)

    code = str(body.get("code", "")).strip().upper()
    if not code:
        return JSONResponse({"error": "Dữ liệu không hợp lệ."}, status_code=422)

    row = db.query(ExternalOrderTracking).filter(ExternalOrderTracking.order_code == code).first()
    if not row:
        return JSONResponse({"error": "Không tìm thấy đơn ngoại sàn."}, status_code=404)

    row.is_paid = 1 if body.get("is_paid") else 0
    row.updated_at = datetime.now()
    db.commit()

    return {
        "ok": True,
        "item": serialize_external_order(row),
    }

@app.post("/api/order-info")
async def get_order_info(request: Request, body: dict, db: Session = Depends(get_db)):
    order_code = body.get("order_code", "").strip()
    key        = body.get("key", "").strip()
    user_id    = request.headers.get('X-User-ID', '')

    if not order_code or not key:
        return JSONResponse({"error": "Thiếu mã đơn hàng hoặc key."}, status_code=400)

    if key not in VALID_KEYS:
        return JSONResponse({"error": "Key không hợp lệ."}, status_code=403)

    if key not in UNLIMITED_KEYS and VALID_KEYS[key] >= KEY_LIMIT:
        return JSONResponse({"error": f"Key đã hết lượt sử dụng ({KEY_LIMIT}/{KEY_LIMIT})."}, status_code=403)

    if len(order_code) < 9:
        return JSONResponse({"error": "Mã đơn hàng không hợp lệ."}, status_code=400)

    VALID_KEYS[key] += 1
    remaining = -1 if key in UNLIMITED_KEYS else KEY_LIMIT - VALID_KEYS[key]

    from datetime import timezone as _tz
    now_vn = datetime.now(_tz(timedelta(hours=7))).strftime("%d/%m/%Y %H:%M")
    if key not in KEY_HISTORY:
        KEY_HISTORY[key] = []
    KEY_HISTORY[key].append({"order_code": order_code, "time": now_vn})

    input_id = order_code[2:9]
    
    url = f"https://ec.megaads.vn/service/inoutput/find-promotion-codes-api?inoutputId={input_id}"
    session = "eyJpdiI6ImIra2pmWitCVVRRTlp2K3pRUUZOZ1E9PSIsInZhbHVlIjoibXpYaFhkQmVZU1VMRFRKWWhEcXRCdnBFSWdycVNzNFlSVHpGWjVYT0hTVDFpdlErVWxDSWhEaVdcL3JyT2RvSjZIcDNkMVJSYTllZDJMMTlsR2ZIQ3BnPT0iLCJtYWMiOiI2MDc2MTFlNDg0MTg4M2IyNDBiNDAzMDE4ZWE0MTk0ZTFkNDdlNGU3MjQ0ZjA3ODFkYTlkYzZiMjcyOTEyMzNmIn0%3D"
    
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            res = await client.get(url, headers={
                "Accept": "application/json, text/plain, */*",
                "platform": "ios",
                "Cookie": f"laravel_session={session}",
                "User-Agent": "chiakiApp/3.6.2"
            })
        data = res.json()

        d = data.get("result") or data.get("data") or {}
        if isinstance(d, list):
            d = d[0] if d else {}
        def g(*keys):
            for k in keys:
                v = d.get(k)
                if v: return str(v)
            return "—"

        phone = next((x for x in d.get("search", "").split() if x.isdigit() and len(x) >= 9), "—")

        payment_type   = d.get("payment_type", "")
        prepaid_amount = d.get("prepaid_amount")
        is_approved    = d.get("is_approved_prepaid", "0")
        prepaid_time   = d.get("prepaid_time")

        if payment_type in ("home", "cod"):
            payment_status = "💵 COD — Thu tiền khi nhận hàng"
        elif payment_type in ("atm", "online", "bank") and is_approved == "1":
            payment_status = f"✅ Đã thanh toán online ({payment_type.upper()})"
            if prepaid_time:
                payment_status += f" lúc {prepaid_time}"
        elif payment_type in ("atm", "online", "bank") and is_approved != "1":
            payment_status = f"⏳ Chờ xác nhận thanh toán ({payment_type.upper()})"
        else:
            payment_status = f"— ({payment_type or 'Không rõ'})"
        db_order = db.query(Order).filter(
        Order.order_code.like(f"%_{order_code}")
            ).first()
        hide_order = should_hide_order(db_order, user_id)
        db_product   = db_order.product   if db_order else "—"
        shop_id_from_api = g("store_code", "creator_name")
        db_shop_name = (
            SHOP_NAME_MAP.get(shop_id_from_api)
            or (db_order.shop_name if db_order else None)
            or shop_id_from_api
        )
        db_total     = f"{int(db_order.total):,} đ".replace(",", ".") if db_order and db_order.total else "—"
        url_history_parsed = []
        try:
            meta_raw = d.get("meta_data", "{}")
            meta = _json.loads(meta_raw) if isinstance(meta_raw, str) else (meta_raw or {})
            uh = meta.get("url_history", {})
            if isinstance(uh, dict):
                url_history_parsed = [v for _, v in sorted(uh.items(), key=lambda x: int(x[0]))]
            elif isinstance(uh, list):
                url_history_parsed = uh
        except Exception:
            meta = {}
            url_history_parsed = []
        source_from = meta.get("meta_tracking", {}).get("from", "") or g("from") or ""

        effective_shop_id = str((db_order.shop_id if db_order and db_order.shop_id else shop_id_from_api) or "").strip()
        if key == "PHONE-KEY-PHUONG2000" and effective_shop_id not in PHONE_KEY_PHUONG_ALLOWED_SHOPS:
            return JSONResponse({"error": "Không có thông tin đơn hàng."}, status_code=404)

        if hide_order:
            return JSONResponse({"error": "Không tìm thấy đơn hàng hoặc bạn không có quyền xem đơn này."}, status_code=404)

        return {
    "order_code":           g("code"),
    "status":               g("status"),
    "shop_name":            db_shop_name,
    "order_date":           g("verified_time", "create_time"),
    "customer_name":        g("related_user_name", "receiver_name"),
    "customer_id":          meta.get("customer_id") or d.get("related_user_id"),
    "phone":                phone,
    "email":                g("email_id"),
    "address":              g("delivery_address"),
    "source":               g("source", "from"),
    "source_from":          source_from,
    "payment":              payment_status,
    "prepaid_amount":       db_total,
    "shipping_code":        g("shipping_code"),
    "delivery_status":      g("delivery_status"),
    "delivery_location_id": g("delivery_location_id"),
    "district_delivery_id": g("district_delivery_id"),
    "commune_delivery_id":  g("commune_delivery_id"),
    "shipper_receive_time": g("shipper_receive_time"),
    "product":              db_product,
    "quantity":             d.get("quantity") or (db_order.quantity if db_order else None),
    "url_history":          url_history_parsed,
    "remaining":            remaining,
}
    except Exception as e:
        VALID_KEYS[key] -= 1
        return JSONResponse({"error": f"Lỗi khi gọi API: {str(e)}"}, status_code=500)

@app.post("/api/order-info/check-key")
async def check_key(body: dict):
    key = body.get("key", "").strip()
    if key not in VALID_KEYS:
        return JSONResponse({"error": "Key không hợp lệ."}, status_code=403)
    used = VALID_KEYS[key]
    is_unlimited = key in UNLIMITED_KEYS
    return {
        "used": used,
        "remaining": -1 if is_unlimited else KEY_LIMIT - used,
        "limit": -1 if is_unlimited else KEY_LIMIT,
        "unlimited": is_unlimited
}
@app.get("/api/order-info/history")
async def get_key_history(request: Request):
    user_id = request.headers.get('X-User-ID', '')
    if user_id != 'Chang2000':
        return JSONResponse({"error": "Không có quyền."}, status_code=403)

    result = []
    for key, logs in KEY_HISTORY.items():
        if logs:
            result.append({
                "key":     key,
                "used":    VALID_KEYS.get(key, 0),
                "limit":   KEY_LIMIT,
                "history": logs
            })
    return {"data": result, "total_queries": sum(len(v) for v in KEY_HISTORY.values())}
@app.post("/api/auth/login-log")
async def login_log(body: dict, request: Request):
    key    = body.get("key", "").strip()
    event  = body.get("event", "login")  # "login" hoặc "logout"
    
    from datetime import timezone as _tz
    now_vn = datetime.now(_tz(timedelta(hours=7))).strftime("%H:%M:%S ngày %d/%m/%Y")
    
    LOGIN_HISTORY.append({
        "key":   key,
        "event": event,
        "time":  now_vn,
    })
    return {"ok": True}

@app.get("/api/auth/login-history")
async def get_login_history(request: Request):
    user_id = request.headers.get('X-User-ID', '')
    if user_id != 'Chang2000':
        return JSONResponse({"error": "Không có quyền."}, status_code=403)
    return {"data": LOGIN_HISTORY, "total": len(LOGIN_HISTORY)}
@app.get("/api/export-order/{order_code}")
async def export_order(order_code: str, request: Request, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # Tìm đơn trong DB — dùng SQLAlchemy như các endpoint khác
    order = db.query(Order).filter(Order.order_code == order_code).first()

    if not order:
        return JSONResponse({"error": "Không tìm thấy đơn hàng"}, status_code=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "Đơn hàng"

    headers = [
        ("Mã đơn",        order.order_code or ""),
        ("Ngày tạo",      order.order_date or ""),
        ("Gian hàng",     order.shop_name or ""),
        ("Tên khách",     order.customer_name or order.buyer_name or ""),
        ("Số điện thoại", order.phone or ""),
        ("Địa chỉ",       order.address or ""),
        ("Sản phẩm",      order.product or ""),
        ("Số lượng",      order.quantity or ""),
        ("Tổng tiền",     order.total or ""),
        ("Trạng thái",    order.status or ""),
    ]

    for i, (label, value) in enumerate(headers, start=1):
        c_label = ws.cell(row=i, column=1, value=label)
        c_label.font = Font(bold=True, color="5B4B8A", size=11)
        c_label.fill = PatternFill("solid", fgColor="EDE8FF")
        c_label.alignment = Alignment(horizontal="left", vertical="center")

        c_val = ws.cell(row=i, column=2, value=str(value) if value else "")
        c_val.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"don-hang-{order_code}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.post("/api/auth/verify-id")
async def verify_id(body: dict):
    VALID_IDS = {
        "LOGIN-KEY-PHUONG2000": {"hours": 24,           "label": "Phương"},
        "LOGIN-KEY-CHANGTESTUSER":    {"hours": 9999999999,   "label": "Hoàng"},
        "Chang2000":    {"hours": 9999999999,   "label": "Hoàng"},
        "unlimited_id": {"hours": 9999999999, "label": "Unlimited"},
    }
    user_id = body.get("id", "").strip()
    info = VALID_IDS.get(user_id)
    if not info:
        return JSONResponse({"error": "ID không hợp lệ."}, status_code=403)
    
    import time
    first_entry = body.get("firstEntry")  # ms từ frontend
    if not first_entry:
        first_entry = int(time.time() * 1000)
    
    exp_ms = first_entry + info["hours"] * 3600000
    if int(time.time() * 1000) > exp_ms:
        return JSONResponse({"error": "ID đã hết hạn."}, status_code=403)
    
    return {"ok": True, "label": info["label"], "expMs": exp_ms, "firstEntry": first_entry}
@app.get("/api/orders/mien-bac")
async def get_mien_bac_orders(request: Request, db: Session = Depends(get_db)):
    user_id = request.headers.get("X-User-ID", "")
    keywords = [
        "hải phòng", "hai phong",
        "bắc giang", "bac giang", "bắc kạn", "bac kan",
        "cao bằng", "cao bang", "hà giang", "ha giang",
        "lạng sơn", "lang son", "phú thọ", "phu tho",
        "quảng ninh", "quang ninh", "thái nguyên", "thai nguyen",
        "tuyên quang", "tuyen quang", "điện biên", "dien bien",
        "hòa bình", "hoa binh", "lai châu", "lai chau",
        "lào cai", "lao cai", "sơn la", "son la",
        "yên bái", "yen bai", "bắc ninh", "bac ninh",
        "hà nam", "ha nam", "hải dương", "hai duong",
        "hưng yên", "hung yen", "nam định", "nam dinh",
        "ninh bình", "ninh binh", "thái bình", "thai binh",
        "vĩnh phúc", "vinh phuc"
    ]
    filters = [func.lower(Order.address).contains(kw.lower()) for kw in keywords]
    q = db.query(Order).filter(or_(*filters))
    if not is_full_access_user(user_id):
        q = q.filter(~Order.shop_id.in_(SENSITIVE_SHOPS)).filter(
            or_(Order.total == None, Order.total < SENSITIVE_TOTAL_THRESHOLD)
        )
    orders = q.order_by(Order.order_date.desc()).all()
    return [serialize_order(o) for o in orders]

@app.get("/api/shops-list")
def get_shops_list():
    shops = get_shops_map()
    result = []
    for shop_id, (shop_url, shop_name) in shops.items():
        result.append({
            "shop_id": shop_id,
            "shop_name": shop_name,
            "shop_url": shop_url
        })
    return sorted(result, key=lambda x: x["shop_name"])

@app.get("/api/sync-config")
def get_sync_config():
    return {
        "seller_id": SELLER_ID,
        "seller_token": SELLER_TOKEN,
    }

@app.post("/api/sync")
async def sync_now(body: dict, db: Session = Depends(get_db)):
    shop_id     = body.get("shop_id", "").strip()
    cf_chl_tk   = body.get("cf_chl_tk", "").strip()
    cf_clearance = body.get("cf_clearance", "").strip()

    if not shop_id:
        return JSONResponse({"error": "Thiếu shop_id"}, status_code=400)
    if not cf_chl_tk or not cf_clearance:
        return JSONResponse({"error": "Thiếu cf_chl_tk hoặc cf_clearance"}, status_code=400)

    shops = get_shops_map()
    if shop_id not in shops:
        return JSONResponse({"error": f"Không tìm thấy shop {shop_id}"}, status_code=404)

    shop_url, shop_name = shops[shop_id]

    try:
        synced = await sync_shop(shop_id, shop_url, shop_name, db,
                                 cf_chl_tk=cf_chl_tk, cf_clearance=cf_clearance)
        return {"ok": True, "shop_id": shop_id, "shop_name": shop_name, "synced": synced}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=502)

@app.post("/api/sync-upload")
async def sync_upload(
    shop_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        shops = get_shops_map()
        if shop_id not in shops:
            return JSONResponse({"error": f"Không tìm thấy shop {shop_id}"}, status_code=404)

        shop_url, shop_name = shops[shop_id]
        content = await file.read()

        # Kiểm tra có phải Excel không (magic bytes PK = xlsx)
        if len(content) < 100:
            return JSONResponse({"error": "File quá nhỏ, không hợp lệ"}, status_code=422)

        orders = parse_excel(content, shop_id, shop_name)
        if orders is None:
            return JSONResponse({"error": "Không đọc được dữ liệu từ file Excel. Kiểm tra lại cột header."}, status_code=422)

        deleted = db.query(Order).filter(Order.shop_id == shop_id).delete()
        for o in orders:
            db.add(Order(**o))
        db.commit()

        meta = db.query(ShopMeta).filter(ShopMeta.shop_id == shop_id).first()
        if meta:
            meta.shop_name   = shop_name
            meta.last_sync   = datetime.now()
            meta.order_count = len(orders)
        else:
            db.add(ShopMeta(
                shop_id=shop_id, shop_name=shop_name,
                shop_url=shop_url, last_sync=datetime.now(),
                order_count=len(orders)
            ))
        db.commit()

        return {
            "ok": True,
            "shop_id": shop_id,
            "shop_name": shop_name,
            "synced": len(orders),
            "deleted": deleted
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)
